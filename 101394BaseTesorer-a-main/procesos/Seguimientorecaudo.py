# -*- coding: utf-8 -*-
import streamlit as st
import zipfile
import io
import pandas as pd
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOMBRES_SALIDA = ["consolidadobancos", "flujotesoreria", "seguimientorecaudo"]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

HOJAS_MESES = {
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
}

HEADER_KEYWORDS = {
    "cuenta", "fecha", "valor", "debito", "credito", "saldo",
    "descripcion", "concepto", "banco", "sucursal", "tipo", "monto",
    "importe", "transaccion", "movimiento", "referencia", "categ",
    "cod", "dcred", "ddeb", "vlr",
}

KEYWORDS_CATEGORIA = ["categoria", "tipo", "tipo de movimiento", "tipo movimiento", "clase", "categ"]
KEYWORDS_CONCEPTO  = ["concepto", "descripcion", "description", "detalle", "referencia", "movimiento"]
KEYWORDS_VALOR     = ["vlr flujo", "valor dc", "valor d/c", "vlr", "valor", "importe", "monto",
                      "credito", "creditos", "dcred"]
KEYWORDS_FECHA     = ["fecha", "date", "fecha movimiento", "fecha valor", "fecha transaccion", "fec"]

RECAUDO_KEYWORDS = [
    "recaudo", "ventas", "recaudo ventas", "recaudo de ventas",
    "consignacion", "consignaciones", "consig",
]


def normalize(text):
    text = str(text).lower().strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", text)


def es_archivo_salida(nombre):
    n = normalize(nombre)
    return any(s in n for s in NOMBRES_SALIDA)


def es_hoja_mes(nombre_hoja):
    nombre_norm = normalize(str(nombre_hoja))
    return any(mes in nombre_norm for mes in HOJAS_MESES)


def find_column(columns, keywords):
    normalized_cols = {normalize(col): col for col in columns}
    for key in keywords:
        key_norm = normalize(key)
        for col_norm, original_col in normalized_cols.items():
            if key_norm in col_norm or col_norm in key_norm:
                return original_col
    return None


def parse_amount(val):
    """
    '232.000'    → 232000.0   (punto = miles COP)
    '232.000,50' → 232000.5   (punto = miles, coma = decimal)
    '1,234.56'   → 1234.56    (formato americano)
    """
    s = str(val).strip().replace("$", "").replace(" ", "")
    if not s or s.lower() in ("nan", "none", "-", ""):
        return 0.0
    has_dot   = "." in s
    has_comma = "," in s
    try:
        if has_dot and has_comma:
            if s.rfind(".") > s.rfind(","):
                s = s.replace(",", "")            # americano: quitar coma-miles
            else:
                s = s.replace(".", "").replace(",", ".")  # COP: quitar punto-miles, coma→punto
        elif has_dot:
            parts = s.split(".")
            if len(parts) > 1 and all(len(p) == 3 for p in parts[1:]):
                s = s.replace(".", "")            # punto = miles
        elif has_comma:
            parts = s.split(",")
            if len(parts) > 1 and all(len(p) == 3 for p in parts[1:]):
                s = s.replace(",", "")            # coma = miles
            else:
                s = s.replace(",", ".")           # coma = decimal
        return float(s)
    except Exception:
        return 0.0


def clean_money(series):
    return series.apply(parse_amount)


def score_header_row(row):
    cells = [normalize(str(x)) for x in row if str(x).strip().lower() not in ("", "nan")]
    return sum(1 for cell in cells if any(kw in cell for kw in HEADER_KEYWORDS))


def find_best_header_row(df_raw, max_search=25):
    """
    Devuelve el índice de la fila con MÁS coincidencias de palabras clave
    bancarias — no la primera que supere un umbral.
    """
    best_row, best_score = 0, 0
    for i in range(min(max_search, len(df_raw))):
        s = score_header_row(df_raw.iloc[i])
        if s > best_score:
            best_score = s
            best_row = i
    return best_row


def read_real_excel(file):
    """
    Lee SOLO las hojas de meses (ENERO, FEBRERO, MARZO…).
    Omite hojas de reconciliación (Rec.Ene, Rec.Feb…).
    """
    try:
        xl = pd.ExcelFile(file)
    except Exception as e:
        return pd.DataFrame(), [], str(e)

    all_dfs, hojas_leidas, detalles = [], [], []

    for sheet_name in xl.sheet_names:
        if not es_hoja_mes(sheet_name):
            detalles.append(f"  ⬜ Omitida: '{sheet_name}'")
            continue
        try:
            df_raw = xl.parse(sheet_name, header=None)
            if df_raw.empty or len(df_raw) < 2:
                detalles.append(f"  ⬛ '{sheet_name}': vacía o muy corta")
                continue

            header_row = find_best_header_row(df_raw)
            df = xl.parse(sheet_name, header=header_row)
            df = df.dropna(how="all")

            # Eliminar filas duplicadas de encabezado
            cols_norm = set(normalize(str(c)) for c in df.columns)
            def es_fila_header(row):
                vals = [normalize(str(v)) for v in row if str(v).strip().lower() not in ("", "nan")]
                return sum(1 for v in vals if v in cols_norm) >= 3
            df = df[~df.apply(es_fila_header, axis=1)].dropna(how="all")

            if not df.empty:
                all_dfs.append(df)
                hojas_leidas.append(sheet_name)
                detalles.append(
                    f"  ✅ '{sheet_name}': header fila {header_row}, "
                    f"{len(df)} filas, cols: {list(df.columns[:6])}"
                )
            else:
                detalles.append(f"  ⚠️ '{sheet_name}': sin datos (header={header_row})")
        except Exception as ex:
            detalles.append(f"  ❌ '{sheet_name}': {ex}")

    if not all_dfs:
        return pd.DataFrame(), hojas_leidas, "\n".join(detalles)

    return pd.concat(all_dfs, ignore_index=True, sort=False), hojas_leidas, "\n".join(detalles)


def cargar_excels(uploads):
    result = []
    for upload in uploads:
        name = upload.name
        if es_archivo_salida(name):
            continue
        data = upload.read()
        if name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for fname in z.namelist():
                    if fname.lower().endswith((".xlsx", ".xls")) and not fname.startswith("__"):
                        base = fname.split("/")[-1]
                        if not es_archivo_salida(base):
                            result.append((base, z.read(fname)))
        elif name.lower().endswith((".xlsx", ".xls")):
            result.append((name, data))
    return result


def standardize_df(df, filename):
    cols = df.columns.tolist()
    n    = len(df)

    categoria_col = find_column(cols, KEYWORDS_CATEGORIA)
    concepto_col  = find_column(cols, KEYWORDS_CONCEPTO)
    valor_col     = find_column(cols, KEYWORDS_VALOR)
    fecha_col     = find_column(cols, KEYWORDS_FECHA)

    categoria = df[categoria_col].fillna("") if categoria_col else pd.Series([""] * n)
    concepto  = df[concepto_col].fillna("")  if concepto_col  else pd.Series([""] * n)
    valor     = clean_money(df[valor_col])   if valor_col     else pd.Series([0.0] * n)

    if fecha_col:
        parsed = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=True)
        if hasattr(parsed.dt, "tz") and parsed.dt.tz is not None:
            parsed = parsed.dt.tz_convert(None)
        fecha = parsed.dt.normalize()
    else:
        fecha = pd.Series([pd.NaT] * n)

    banco = str(filename).replace(".xlsx", "").replace(".xls", "")
    return pd.DataFrame({
        "Categoria": categoria.astype(str).values,
        "Concepto":  concepto.astype(str).values,
        "Banco":     banco,
        "Fecha":     fecha.values,
        "Valor":     valor.values,
    }), {"fecha_col": fecha_col, "valor_col": valor_col,
         "categoria_col": categoria_col, "concepto_col": concepto_col}


def es_recaudo(concepto_str):
    norm = normalize(str(concepto_str))
    return any(normalize(kw) in norm for kw in RECAUDO_KEYWORDS)


HEADER_COLOR   = "1F4E79"
WEEKEND_BG     = "FCE4D6"
ALT_BG         = "EBF3FB"
GRAND_TOTAL_BG = "163755"


def safe_val(v):
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, float) and pd.isna(v):
        return 0.0
    return v


def write_daily_sheet(ws, pivot, banco_cols):
    all_cols   = ["Fecha"] + banco_cols + ["Total Diario", "Total Acumulado Mes"]
    money_cols = banco_cols + ["Total Diario", "Total Acumulado Mes"]

    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    alt_fill    = PatternFill("solid", fgColor=ALT_BG)
    wknd_fill   = PatternFill("solid", fgColor=WEEKEND_BG)
    grand_fill  = PatternFill("solid", fgColor=GRAND_TOTAL_BG)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    grand_font  = Font(bold=True, color="FFFFFF", size=11)

    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = header_fill; cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, (_, row) in enumerate(pivot.iterrows(), start=2):
        fecha_val  = row["Fecha"]
        is_weekend = isinstance(fecha_val, pd.Timestamp) and fecha_val.weekday() >= 5
        row_fill   = wknd_fill if is_weekend else (alt_fill if ri % 2 == 0 else None)
        for ci, col_name in enumerate(all_cols, start=1):
            v    = safe_val(row[col_name])
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col_name == "Fecha":
                if hasattr(v, "strftime"):
                    cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            elif col_name in money_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    tr     = ws.max_row + 1
    totals = {"Fecha": "TOTAL"}
    for b in banco_cols:
        totals[b] = float(pivot[b].fillna(0).sum())
    totals["Total Diario"]        = float(pivot["Total Diario"].fillna(0).sum())
    totals["Total Acumulado Mes"] = float(pivot["Total Acumulado Mes"].fillna(0).iloc[-1]) if not pivot.empty else 0.0

    for ci, col_name in enumerate(all_cols, start=1):
        v    = totals.get(col_name, 0)
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.fill = grand_fill; cell.font = grand_font; cell.border = border
        if col_name in money_cols:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    for ci in range(1, len(all_cols) + 1):
        maxlen = max(len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 4, 30)

    ws.freeze_panes = "B2"


def build_pivot_mes(df_mes, banco_cols):
    agg = df_mes.groupby(["Fecha", "Banco"])["Valor"].sum().reset_index()
    if agg.empty:
        return pd.DataFrame(columns=["Fecha"] + banco_cols + ["Total Diario", "Total Acumulado Mes"])

    pivot = agg.pivot_table(
        index="Fecha", columns="Banco", values="Valor",
        fill_value=0.0, aggfunc="sum",
    ).reset_index()
    pivot.columns.name = None

    for b in banco_cols:
        if b not in pivot.columns:
            pivot[b] = 0.0

    pivot = pivot.sort_values("Fecha").reset_index(drop=True)
    pivot = pivot[["Fecha"] + banco_cols]
    pivot["Total Diario"]        = pivot[banco_cols].sum(axis=1)
    pivot["Total Acumulado Mes"] = pivot["Total Diario"].cumsum()
    return pivot


def build_excel(data_por_mes, banco_cols):
    wb = Workbook()
    wb.remove(wb.active)

    for mes_label, df_mes in data_por_mes.items():
        pivot = build_pivot_mes(df_mes, banco_cols)
        ws    = wb.create_sheet(title=mes_label[:31])
        write_daily_sheet(ws, pivot, banco_cols)

    df_todo     = pd.concat(list(data_por_mes.values()), ignore_index=True)
    pivot_total = build_pivot_mes(df_todo, banco_cols)
    ws_t        = wb.create_sheet(title="Total")
    write_daily_sheet(ws_t, pivot_total, banco_cols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run():
    st.set_page_config(page_title="Seguimiento Recaudo Diario", page_icon="🏦", layout="wide")
    st.title("🏦 Seguimiento Recaudo Diario")
    st.markdown(
        "Sube los archivos Excel de los bancos o un ZIP. "
        "Lee las hojas **ENERO, FEBRERO, MARZO…** y genera una hoja por mes."
    )

    uploads = st.file_uploader(
        "Archivos Excel o ZIP",
        type=["xlsx", "xls", "zip"],
        accept_multiple_files=True,
        key="sr_files",
    )
    if not uploads:
        return

    try:
        archivos = cargar_excels(uploads)
        if not archivos:
            st.warning("No se encontraron archivos de banco válidos.")
            return

        st.info(f"📂 Archivos a procesar: **{len(archivos)}**")

        all_data, report, diag_lines = [], [], []

        for nombre, data in archivos:
            diag_lines.append(f"\n{nombre}")
            try:
                df, hojas_leidas, detalles = read_real_excel(io.BytesIO(data))
                diag_lines.append(detalles)

                if df.empty:
                    report.append({"Archivo": nombre, "Hojas leídas": "Ninguna",
                                   "Total filas": 0, "Filas recaudo": 0, "Estado": "Sin hojas de mes"})
                    continue

                clean_df, col_info = standardize_df(df, nombre)
                diag_lines.append(
                    f"  fecha={col_info['fecha_col']} | valor={col_info['valor_col']} | "
                    f"categoria={col_info['categoria_col']} | concepto={col_info['concepto_col']}"
                )

                fechas_validas = clean_df["Fecha"].dropna()
                if not fechas_validas.empty:
                    dist = fechas_validas.dt.to_period("M").value_counts().sort_index()
                    diag_lines.append("  Fechas por mes: " + ", ".join(f"{p}: {n}" for p, n in dist.items()))
                else:
                    diag_lines.append("  ⚠️ Sin fechas válidas")

                mask_cat = clean_df["Categoria"].apply(
                    lambda x: normalize(x) in {"ingreso", "ingresos", "credito", "creditos"}
                )
                mask_con = clean_df["Concepto"].apply(es_recaudo)
                filtered = clean_df[mask_cat & mask_con].copy()
                if filtered.empty:
                    filtered = clean_df[mask_con].copy()

                report.append({
                    "Archivo":       nombre,
                    "Hojas leídas":  ", ".join(hojas_leidas),
                    "Total filas":   len(clean_df),
                    "Filas recaudo": len(filtered),
                    "Estado":        "OK" if not filtered.empty else "Sin filas de recaudo",
                })
                if not filtered.empty:
                    all_data.append(filtered)

            except Exception as e:
                report.append({"Archivo": nombre, "Hojas leídas": "", "Total filas": 0,
                                "Filas recaudo": 0, "Estado": str(e)})
                diag_lines.append(f"  ❌ {e}")

        with st.expander("📋 Reporte de lectura"):
            st.dataframe(pd.DataFrame(report), use_container_width=True)

        with st.expander("🔍 Diagnóstico detallado (hojas, columnas, fechas)"):
            st.text("\n".join(diag_lines))

        if not all_data:
            st.warning("⚠️ No se encontraron filas de recaudo. Revisa el diagnóstico.")
            return

        final_df = pd.concat(all_data, ignore_index=True)
        final_df = final_df[final_df["Fecha"].notna()].copy()

        if final_df.empty:
            st.warning("No hay registros con fecha válida.")
            return

        final_df["Mes_num"]   = final_df["Fecha"].dt.month
        final_df["Anio"]      = final_df["Fecha"].dt.year
        final_df["Mes"]       = final_df["Mes_num"].map(MESES_ES)
        final_df["Mes_label"] = final_df.apply(lambda r: f"{r['Mes']} {int(r['Anio'])}", axis=1)

        orden = (final_df[["Anio", "Mes_num", "Mes_label"]]
                 .drop_duplicates().sort_values(["Anio", "Mes_num"]))
        meses_ordenados = list(orden["Mes_label"])
        banco_cols      = sorted(final_df["Banco"].unique().tolist())

        data_por_mes = {
            mes: final_df[final_df["Mes_label"] == mes]
            for mes in meses_ordenados
            if not final_df[final_df["Mes_label"] == mes].empty
        }

        c1, c2, c3 = st.columns(3)
        c1.metric("Meses encontrados", len(data_por_mes))
        c2.metric("Bancos", len(banco_cols))
        c3.metric("Total recaudo", "${:,.0f}".format(final_df["Valor"].sum()))

        st.markdown("---")
        tabs = st.tabs(list(data_por_mes.keys()) + ["📊 Total"])
        for tab, (mes_label, df_mes) in zip(tabs, list(data_por_mes.items()) + [("Total", final_df)]):
            with tab:
                st.dataframe(build_pivot_mes(df_mes, banco_cols), use_container_width=True, hide_index=True)

        st.markdown("---")
        st.download_button(
            label="⬇️ Descargar Seguimiento Recaudo Diario (.xlsx)",
            data=build_excel(data_por_mes, banco_cols),
            file_name="Seguimiento_Recaudo_Diario.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error: {e}")
        st.exception(e)


if __name__ == "__main__":
    run()