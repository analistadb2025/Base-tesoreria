# -*- coding: utf-8 -*-
import streamlit as st
import zipfile
import io
import pandas as pd
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOMBRES_SALIDA = ["consolidadobancos", "flujotesoreria", "seguimientorecaudo"]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

KEYWORDS_CATEGORIA = ["categoria", "tipo", "tipo de movimiento", "tipo movimiento", "clase"]
KEYWORDS_CONCEPTO  = ["concepto", "descripcion", "description", "detalle", "referencia", "movimiento"]
KEYWORDS_VALOR     = ["vlr flujo", "valor dc", "valor d/c", "vlr", "valor", "importe", "monto"]
KEYWORDS_FECHA     = ["fecha", "date", "fecha movimiento", "fecha valor", "fecha transaccion", "fec"]

VALORES_INGRESO = {"ingreso", "ingresos", "credito", "creditos", "entrada", "entradas"}
VALORES_EGRESO  = {"egreso", "egresos", "gasto", "gastos", "salida", "salidas", "debito", "debitos", "pago", "pagos"}


def normalize(text):
    text = str(text).lower().strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", text)


def es_archivo_salida(nombre):
    n = normalize(nombre)
    return any(s in n for s in NOMBRES_SALIDA)


def find_column(columns, keywords):
    normalized_cols = {normalize(col): col for col in columns}
    for key in keywords:
        key_norm = normalize(key)
        for col_norm, original_col in normalized_cols.items():
            if key_norm in col_norm or col_norm in key_norm:
                return original_col
    return None


def clean_money(series):
    return (
        series.astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0.0)
    )


def is_header_row(row):
    text_cells = [str(x).strip() for x in row if isinstance(x, str) and len(str(x).strip()) > 0]
    return len([t for t in text_cells if len(t) < 30]) >= 4


def read_real_excel(file):
    df_raw = pd.read_excel(file, header=None)
    header_row = 0
    for i, row in df_raw.iterrows():
        if is_header_row(row):
            header_row = i
            break
    return pd.read_excel(file, header=header_row)


def cargar_excels(uploads):
    result   = []
    omitidos = []
    for upload in uploads:
        name = upload.name
        if es_archivo_salida(name):
            omitidos.append(name)
            upload.read()
            continue
        data = upload.read()
        if name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for fname in z.namelist():
                    if fname.lower().endswith((".xlsx", ".xls")) and not fname.startswith("__"):
                        base = fname.split("/")[-1]
                        if not es_archivo_salida(base):
                            result.append((base, z.read(fname)))
                        else:
                            omitidos.append(base)
        elif name.lower().endswith((".xlsx", ".xls")):
            result.append((name, data))
    return result, omitidos


def standardize_df(df, filename):
    cols = df.columns.tolist()
    n = len(df)
    categoria_col = find_column(cols, KEYWORDS_CATEGORIA)
    concepto_col  = find_column(cols, KEYWORDS_CONCEPTO)
    valor_col     = find_column(cols, KEYWORDS_VALOR)
    fecha_col     = find_column(cols, KEYWORDS_FECHA)

    categoria = df[categoria_col].fillna("Sin categoria") if categoria_col else pd.Series(["Sin categoria"] * n)
    concepto  = df[concepto_col].fillna("Sin concepto") if concepto_col else pd.Series(["Sin concepto"] * n)
    valor     = clean_money(df[valor_col]) if valor_col else pd.Series([0.0] * n)

    if fecha_col:
        parsed = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=True)
        if hasattr(parsed.dt, "tz") and parsed.dt.tz is not None:
            parsed = parsed.dt.tz_convert(None)
        fecha = parsed.dt.normalize()
    else:
        fecha = pd.Series([pd.NaT] * n)

    banco = str(filename).replace(".xlsx", "").replace(".xls", "")
    return pd.DataFrame({
        "Categoria": categoria.astype(str).values,
        "Concepto":  concepto.astype(str).values,
        "Banco":     banco,
        "Fecha":     fecha.values,
        "Valor":     valor.values,
    }), {"categoria_col": categoria_col, "concepto_col": concepto_col,
         "valor_col": valor_col, "fecha_col": fecha_col}


def es_ingreso(cat_str):
    return normalize(cat_str) in VALORES_INGRESO


def es_egreso(cat_str):
    return normalize(cat_str) in VALORES_EGRESO


def safe_float(v):
    if isinstance(v, float) and pd.isna(v):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def build_excel_flujo(final_pivot, mes_cols):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo de Tesoreria"

    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    ing_fill    = PatternFill("solid", fgColor="E2EFDA")
    egr_fill    = PatternFill("solid", fgColor="FCE4D6")
    flujo_fill  = PatternFill("solid", fgColor="BDD7EE")
    saldo_fill  = PatternFill("solid", fgColor="9DC3E6")
    sub_ing_f   = PatternFill("solid", fgColor="C6EFCE")
    sub_egr_f   = PatternFill("solid", fgColor="FFCCCC")
    sub_fill    = PatternFill("solid", fgColor="D9D9D9")
    alt_fill    = PatternFill("solid", fgColor="F2F7FB")
    bold_font   = Font(bold=True, size=10)
    flujo_font  = Font(bold=True, size=11)

    all_cols   = list(final_pivot.columns)
    money_cols = mes_cols + ["Total"]

    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, (_, row) in enumerate(final_pivot.iterrows(), start=2):
        cat = str(row.get("Categoria", "")).strip().upper()
        if cat == "FLUJO NETO":
            row_fill, row_font = flujo_fill, flujo_font
        elif cat == "SALDO ACUMULADO":
            row_fill, row_font = saldo_fill, flujo_font
        elif cat == "TOTAL INGRESOS":
            row_fill, row_font = sub_ing_f, bold_font
        elif cat == "TOTAL EGRESOS":
            row_fill, row_font = sub_egr_f, bold_font
        elif cat == "TOTAL OTROS":
            row_fill, row_font = sub_fill, bold_font
        elif es_ingreso(str(row.get("Categoria", ""))):
            row_fill, row_font = ing_fill, Font(size=10)
        elif es_egreso(str(row.get("Categoria", ""))):
            row_fill, row_font = egr_fill, Font(size=10)
        elif ri % 2 == 0:
            row_fill, row_font = alt_fill, Font(size=10)
        else:
            row_fill, row_font = None, Font(size=10)

        for ci, col_name in enumerate(all_cols, start=1):
            raw  = row.get(col_name, 0)
            v    = safe_float(raw) if col_name in money_cols else raw
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.font   = row_font
            if row_fill:
                cell.fill = row_fill
            if col_name in money_cols and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    for ci in range(1, len(all_cols) + 1):
        maxlen = max(
            len(str(ws.cell(row=r, column=ci).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 4, 40)

    ws.freeze_panes = "C2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run():
    st.title("Flujo de Tesoreria")
    st.markdown(
        "Sube los archivos Excel de los bancos o un ZIP. "
        "El sistema agrupa por mes (columnas) y concepto (filas)."
    )

    uploads = st.file_uploader(
        "Archivos Excel o ZIP",
        type=["xlsx", "xls", "zip"],
        accept_multiple_files=True,
        key="ft_files",
    )

    if not uploads:
        return

    try:
        archivos, omitidos = cargar_excels(uploads)

        if omitidos:
            st.info("Omitidos (reportes anteriores): {}".format(", ".join(omitidos)))

        if not archivos:
            st.warning("No se encontraron archivos de banco validos.")
            return

        st.info("Archivos a procesar: {}".format(len(archivos)))

        all_data = []
        report   = []

        for nombre, data in archivos:
            try:
                df            = read_real_excel(io.BytesIO(data))
                clean_df, det = standardize_df(df, nombre)
                fechas_ok     = int(clean_df["Fecha"].notna().sum())
                cats_muestra  = clean_df["Categoria"].unique().tolist()[:8]
                report.append({
                    "Archivo":       nombre,
                    "Filas":         len(clean_df),
                    "Col Categoria": det["categoria_col"] or "NO DETECTADA",
                    "Col Valor":     det["valor_col"] or "NO DETECTADA",
                    "Col Fecha":     det["fecha_col"] or "NO DETECTADA",
                    "Filas c/fecha": fechas_ok,
                    "Categorias":    str(cats_muestra),
                    "Estado":        "OK",
                })
                all_data.append(clean_df)
            except Exception as e:
                report.append({"Archivo": nombre, "Filas": 0, "Col Categoria": "ERROR",
                                "Col Valor": "ERROR", "Col Fecha": "ERROR",
                                "Filas c/fecha": 0, "Categorias": "", "Estado": str(e)})

        with st.expander("Diagnostico de lectura", expanded=True):
            st.dataframe(pd.DataFrame(report), use_container_width=True)
            st.caption("Si 'Col Categoria' dice NO DETECTADA, comparte una captura del Excel del banco.")

        if not all_data:
            st.error("Ningun archivo aporto datos.")
            return

        final_df = pd.concat(all_data, ignore_index=True)
        df_f     = final_df[final_df["Fecha"].notna()].copy()
        n_sin    = int(final_df["Fecha"].isna().sum())

        if n_sin > 0:
            st.warning("{} filas sin fecha seran ignoradas.".format(n_sin))

        if df_f.empty:
            st.error("No se encontro columna de fecha. Revisa el diagnostico.")
            return

        cats = sorted(df_f["Categoria"].unique().tolist())
        with st.expander("Categorias encontradas"):
            st.write(cats)
            st.caption("Ingresos: ingreso/ingresos/credito/entrada. Egresos: egreso/egresos/gasto/gastos/salida/debito/pago.")

        df_f["Mes_num"]   = df_f["Fecha"].dt.month
        df_f["Anio"]      = df_f["Fecha"].dt.year
        df_f["Mes"]       = df_f["Mes_num"].map(MESES_ES)
        df_f["Mes_label"] = df_f.apply(lambda r: "{} {}".format(r["Mes"], int(r["Anio"])), axis=1)

        orden = (
            df_f[["Anio", "Mes_num", "Mes_label"]]
            .drop_duplicates()
            .sort_values(["Anio", "Mes_num"])
        )
        mes_cols = list(orden["Mes_label"])

        grouped = (
            df_f
            .groupby(["Categoria", "Concepto", "Mes_label"])["Valor"]
            .sum()
            .reset_index()
        )

        if grouped.empty:
            st.error("Sin datos tras agrupar.")
            return

        pivot = grouped.pivot_table(
            index=["Categoria", "Concepto"],
            columns="Mes_label",
            values="Valor",
            fill_value=0.0,
            aggfunc="sum",
        ).reset_index()
        pivot.columns.name = None

        for m in mes_cols:
            if m not in pivot.columns:
                pivot[m] = 0.0
        existing = [m for m in mes_cols if m in pivot.columns]
        pivot    = pivot[["Categoria", "Concepto"] + existing]
        pivot["Total"] = pivot[existing].fillna(0).sum(axis=1)

        num_cols    = existing + ["Total"]
        mask_ing    = pivot["Categoria"].apply(es_ingreso)
        mask_egr    = pivot["Categoria"].apply(es_egreso)
        ingresos_df = pivot[mask_ing].copy()
        egresos_df  = pivot[mask_egr].copy()
        otros_df    = pivot[~mask_ing & ~mask_egr].copy()

        st.info("Ingresos: {} | Egresos: {} | Otros: {}".format(
            len(ingresos_df), len(egresos_df), len(otros_df)))

        def totals_row(df, label_cat, label_conc=""):
            row = {c: safe_float(df[c].fillna(0).sum()) for c in num_cols}
            row["Categoria"] = label_cat
            row["Concepto"]  = label_conc
            return pd.DataFrame([row])

        sections = []
        if not ingresos_df.empty:
            sections.append(ingresos_df)
            sections.append(totals_row(ingresos_df, "TOTAL INGRESOS"))
        if not egresos_df.empty:
            sections.append(egresos_df)
            sections.append(totals_row(egresos_df, "TOTAL EGRESOS"))
        if not otros_df.empty:
            sections.append(otros_df)
            sections.append(totals_row(otros_df, "OTROS", "TOTAL OTROS"))

        if not ingresos_df.empty or not egresos_df.empty:
            flujo_row = {"Categoria": "FLUJO NETO", "Concepto": "Ingresos - Egresos"}
            for c in num_cols:
                ing = safe_float(ingresos_df[c].fillna(0).sum()) if not ingresos_df.empty else 0.0
                egr = safe_float(egresos_df[c].fillna(0).sum()) if not egresos_df.empty else 0.0
                flujo_row[c] = ing - egr
            sections.append(pd.DataFrame([flujo_row]))

            saldo_row = {"Categoria": "SALDO ACUMULADO", "Concepto": "Acumulado del periodo"}
            acum = 0.0
            for c in existing:
                acum += flujo_row[c]
                saldo_row[c] = acum
            saldo_row["Total"] = safe_float(flujo_row["Total"])
            sections.append(pd.DataFrame([saldo_row]))

        if not sections:
            st.error("No se generaron secciones. Revisa las categorias.")
            return

        final_pivot = pd.concat(sections, ignore_index=True)
        col_order   = ["Categoria", "Concepto"] + existing + ["Total"]
        final_pivot = final_pivot[col_order]

        st.success("Flujo listo: {} mes(es) | {} conceptos".format(len(existing), len(pivot)))

        df_show = final_pivot.copy()
        for c in existing + ["Total"]:
            df_show[c] = df_show[c].apply(
                lambda x: "${:,.2f}".for **...**

_This response is too long to display in full._