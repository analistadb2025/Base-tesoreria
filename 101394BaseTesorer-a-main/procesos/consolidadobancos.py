import streamlit as st
import zipfile
import io
import pandas as pd
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def normalize(text):
    text = str(text).lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", text)


def find_column(columns, keywords):
    normalized_cols = {normalize(col): col for col in columns}
    for key in keywords:
        key_norm = normalize(key)
        for col_norm, original_col in normalized_cols.items():
            if key_norm in col_norm or col_norm in key_norm:
                return original_col
    return None


def clean_money(series):
    return (
        series.astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0)
    )


def is_header_row(row):
    text_cells = [str(x).strip() for x in row if isinstance(x, str) and len(str(x).strip()) > 0]
    short_texts = [t for t in text_cells if len(t) < 30]
    return len(short_texts) >= 4


def read_real_excel(file):
    df_raw = pd.read_excel(file, header=None)
    header_row = None
    for i, row in df_raw.iterrows():
        if is_header_row(row):
            header_row = i
            break
    if header_row is None:
        header_row = 0
    return pd.read_excel(file, header=header_row), header_row


COLUMN_MAP = {
    "categoria": ["categoria"],
    "concepto": ["concepto"],
    "valor": ["vlr flujo", "valor dc", "valor d/c", "vlr", "valor", "valor de la compra"],
}


def standardize_df(df, filename):
    cols = df.columns.tolist()
    n = len(df)
    categoria_col = find_column(cols, COLUMN_MAP["categoria"])
    concepto_col = find_column(cols, COLUMN_MAP["concepto"])
    valor_col = find_column(cols, COLUMN_MAP["valor"])

    categoria = df[categoria_col] if categoria_col else pd.Series([None] * n)
    concepto = df[concepto_col] if concepto_col else pd.Series([None] * n)
    total = clean_money(df[valor_col]) if valor_col else pd.Series([0] * n)
    banco = filename.split("/")[-1].replace(".xlsx", "")

    return pd.DataFrame({
        "Categoria": categoria.values,
        "Concepto": concepto.values,
        "Banco": banco,
        "Total": total.values,
    })


def build_excel_consolidado(pivot_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado Bancos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    money_fmt = '#,##0.00'

    cols = list(pivot_df.columns)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(pivot_df.itertuples(index=False), start=2):
        is_total = str(row[0]).upper() == "TOTAL" or str(row[1]).upper() == "TOTAL"
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if is_total:
                cell.fill = total_fill
                cell.font = total_font
            col_name = cols[col_idx - 1]
            if col_name not in ("Categoria", "Concepto") and isinstance(value, (int, float)):
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(cols) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run():
    st.title("Consolidado Bancos")
    st.markdown("Suba el ZIP con los Excel de todos los bancos.")

    zip_file = st.file_uploader("Suba el ZIP con los Excel", type="zip", key="cb_zip")

    if zip_file is not None:
        all_data = []
        report = []

        with zipfile.ZipFile(io.BytesIO(zip_file.read())) as z:
            excel_files = [f for f in z.namelist() if f.lower().endswith(".xlsx") and not f.startswith("__")]
            st.info(f"Archivos encontrados: {len(excel_files)}")

            for file in excel_files:
                with z.open(file) as f:
                    try:
                        df, _ = read_real_excel(f)
                        clean_df = standardize_df(df, file)
                        report.append({"Archivo": file, "Filas": len(clean_df), "Estado": "✅ OK"})
                        all_data.append(clean_df)
                    except Exception as e:
                        report.append({"Archivo": file, "Filas": 0, "Estado": f"❌ {e}"})

        with st.expander("Reporte de lectura por archivo"):
            st.dataframe(pd.DataFrame(report), use_container_width=True)

        if all_data:
            final_df = pd.concat(all_data, ignore_index=True)

            pivot_df = (
                final_df
                .groupby(["Categoria", "Concepto", "Banco"], dropna=False)["Total"]
                .sum()
                .reset_index()
                .pivot_table(
                    index=["Categoria", "Concepto"],
                    columns="Banco",
                    values="Total",
                    fill_value=0,
                )
                .reset_index()
            )
            pivot_df.columns.name = None
            banco_cols = [c for c in pivot_df.columns if c not in ("Categoria", "Concepto")]
            pivot_df["Total"] = pivot_df[banco_cols].sum(axis=1)

            totals = {c: pivot_df[c].sum() for c in banco_cols + ["Total"]}
            totals["Categoria"] = "TOTAL"
            totals["Concepto"] = ""
            pivot_df_display = pd.concat([pivot_df, pd.DataFrame([totals])], ignore_index=True)

            st.success(f"Consolidado listo: {len(pivot_df)} conceptos de {len(banco_cols)} banco(s)")
            st.dataframe(pivot_df_display, use_container_width=True)

            excel_buf = build_excel_consolidado(pivot_df_display)
            st.download_button(
                "⬇️ Descargar Consolidado Bancos",
                excel_buf,
                "Consolidado_Bancos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.error("Ningún archivo aportó datos útiles.")