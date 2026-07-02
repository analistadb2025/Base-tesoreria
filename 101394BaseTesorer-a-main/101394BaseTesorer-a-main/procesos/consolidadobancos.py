# -*- coding: utf-8 -*-
import streamlit as st
import zipfile
import io
import pandas as pd
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOMBRES_SALIDA = ["consolidadobancos", "flujotesoreria", "seguimientorecaudo"]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

KEYWORDS_CATEGORIA = ["categoria", "tipo", "tipo de movimiento", "tipo movimiento", "clase"]
KEYWORDS_CONCEPTO  = ["concepto", "descripcion", "description", "detalle", "referencia", "movimiento"]
KEYWORDS_VALOR     = ["vlr flujo", "valor dc", "valor d/c", "vlr", "valor", "importe", "monto"]
KEYWORDS_FECHA     = ["fecha", "date", "fecha movimiento", "fecha valor", "fecha transaccion", "fec"]


def normalize(text):
    text = str(text).lower().strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", text)


def es_archivo_salida(nombre):
    n = normalize(nombre)
    return any(s in n for s in NOMBRES_SALIDA)


def find_column(columns, keywords):
    normalized_cols = {normalize(col): col for col in columns}
    for key in keywords:
        key_norm = normalize(key)
        for col_norm, original_col in normalized_cols.items():
            if key_norm in col_norm or col_norm in key_norm:
                return original_col
    return None


def clean_money(series):
    return (
        series.astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0.0)
    )


def is_header_row(row):
    text_cells = [str(x).strip() for x in row if isinstance(x, str) and len(str(x).strip()) > 0]
    return len([t for t in text_cells if len(t) < 30]) >= 4


def read_real_excel(file):
    df_raw = pd.read_excel(file, header=None)
    header_row = 0
    for i, row in df_raw.iterrows():
        if is_header_row(row):
            header_row = i
            break
    return pd.read_excel(file, header=header_row)


def cargar_excels(uploads):
    result = []
    for upload in uploads:
        name = upload.name
        if es_archivo_salida(name):
            continue
        data = upload.read()
        if name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for fname in z.namelist():
                    if fname.lower().endswith((".xlsx", ".xls")) and not fname.startswith("__"):
                        base = fname.split("/")[-1]
                        if not es_archivo_salida(base):
                            result.append((base, z.read(fname)))
        elif name.lower().endswith((".xlsx", ".xls")):
            result.append((name, data))
    return result


def standardize_df(df, filename):
    cols = df.columns.tolist()
    n = len(df)
    categoria_col = find_column(cols, KEYWORDS_CATEGORIA)
    concepto_col  = find_column(cols, KEYWORDS_CONCEPTO)
    valor_col     = find_column(cols, KEYWORDS_VALOR)
    fecha_col     = find_column(cols, KEYWORDS_FECHA)

    categoria = df[categoria_col].fillna("Sin categoria") if categoria_col else pd.Series(["Sin categoria"] * n)
    concepto  = df[concepto_col].fillna("Sin concepto") if concepto_col else pd.Series(["Sin concepto"] * n)
    valor     = clean_money(df[valor_col]) if valor_col else pd.Series([0.0] * n)

    if fecha_col:
        parsed = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=True)
        if hasattr(parsed.dt, "tz") and parsed.dt.tz is not None:
            parsed = parsed.dt.tz_convert(None)
        fecha = parsed.dt.normalize()
    else:
        fecha = pd.Series([pd.NaT] * n)

    banco = str(filename).replace(".xlsx", "").replace(".xls", "")
    return pd.DataFrame({
        "Categoria": categoria.astype(str).values,
        "Concepto":  concepto.astype(str).values,
        "Banco":     banco,
        "Fecha":     fecha.values,
        "Valor":     valor.values,
    })


def safe_val(v):
    if isinstance(v, float) and pd.isna(v):
        return 0.0
    return v


def build_pivot_concepto(df_mes, banco_cols):
    agg = df_mes.groupby(["Categoria", "Concepto", "Banco"])["Valor"].sum().reset_index()

    if agg.empty:
        return pd.DataFrame(columns=["Categoria", "Concepto"] + banco_cols + ["Total"])

    pivot = agg.pivot_table(
        index=["Categoria", "Concepto"],
        columns="Banco",
        values="Valor",
        fill_value=0.0,
        aggfunc="sum",
    ).reset_index()
    pivot.columns.name = None

    for b in banco_cols:
        if b not in pivot.columns:
            pivot[b] = 0.0

    pivot = pivot[["Categoria", "Concepto"] + banco_cols]
    pivot["Total"] = pivot[banco_cols].sum(axis=1)
    pivot = pivot.sort_values(["Categoria", "Concepto"]).reset_index(drop=True)
    return pivot


def write_concepto_sheet(ws, pivot, banco_cols):
    all_cols   = ["Categoria", "Concepto"] + banco_cols + ["Total"]
    money_cols = banco_cols + ["Total"]

    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill  = PatternFill("solid", fgColor="163755")
    ing_fill    = PatternFill("solid", fgColor="E2EFDA")
    egr_fill    = PatternFill("solid", fgColor="FCE4D6")
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_font  = Font(bold=True, color="FFFFFF", size=11)

    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, (_, row) in enumerate(pivot.iterrows(), start=2):
        cat_norm = normalize(str(row.get("Categoria", "")))
        is_ing   = cat_norm in {"ingreso", "ingresos", "credito", "creditos"}
        is_egr   = cat_norm in {"egreso", "egresos", "gasto", "gastos", "salida", "debito"}
        row_fill = ing_fill if is_ing else (egr_fill if is_egr else (alt_fill if ri % 2 == 0 else None))

        for ci, col_name in enumerate(all_cols, start=1):
            v    = safe_val(row.get(col_name, 0))
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col_name in money_cols and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    tr = ws.max_row + 1
    totals_row = {"Categoria": "TOTAL", "Concepto": ""}
    for b in banco_cols:
        totals_row[b] = float(pivot[b].fillna(0).sum())
    totals_row["Total"] = float(pivot["Total"].fillna(0).sum())

    for ci, col_name in enumerate(all_cols, start=1):
        v    = totals_row.get(col_name, 0)
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.fill   = total_fill
        cell.font   = total_font
        cell.border = border
        if col_name in money_cols and isinstance(v, (int, float)):
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    for ci in range(1, len(all_cols) + 1):
        maxlen = max(
            len(str(ws.cell(row=r, column=ci).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 4, 40)

    ws.freeze_panes = "C2"


def build_excel(data_por_mes, banco_cols):
    wb = Workbook()
    wb.remove(wb.active)

    for mes_label, df_mes in data_por_mes.items():
        pivot = build_pivot_concepto(df_mes, banco_cols)
        ws = wb.create_sheet(title=mes_label[:31])
        write_concepto_sheet(ws, pivot, banco_cols)

    df_todo = pd.concat(list(data_por_mes.values()), ignore_index=True)
    pivot_total = build_pivot_concepto(df_todo, banco_cols)
    ws_t = wb.create_sheet(title="Total")
    write_concepto_sheet(ws_t, pivot_total, banco_cols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run():
    st.title("Consolidado Bancos")
    st.markdown(
        "Sube los archivos Excel de los bancos o un ZIP. "
        "El Excel genera **una hoja por mes** con filas = conceptos y columnas = banco."
    )

    uploads = st.file_uploader(
        "Archivos Excel o ZIP",
        type=["xlsx", "xls", "zip"],
        accept_multiple_files=True,
        key="cb_files",
    )

    if not uploads:
        return

    try:
        archivos = cargar_excels(uploads)
        if not archivos:
            st.warning("No se encontraron archivos de banco validos.")
            return

        st.info("Archivos a procesar: {}".format(len(archivos)))

        all_data = []
        report   = []

        for nombre, data in archivos:
            try:
                df       = read_real_excel(io.BytesIO(data))
                clean_df = standardize_df(df, nombre)
                report.append({"Archivo": nombre, "Filas": len(clean_df), "Estado": "OK"})
                all_data.append(clean_df)
            except Exception as e:
                report.append({"Archivo": nombre, "Filas": 0, "Estado": str(e)})

        with st.expander("Reporte de lectura"):
            st.dataframe(pd.DataFrame(report), use_container_width=True)

        if not all_data:
            st.error("Ningun archivo aporto datos.")
            return

        final_df   = pd.concat(all_data, ignore_index=True)
        banco_cols = sorted(final_df["Banco"].unique().tolist())

        tiene_fechas = final_df["Fecha"].notna().any()
        if tiene_fechas:
            final_df = final_df[final_df["Fecha"].notna()].copy()
            final_df["Mes_num"]   = final_df["Fecha"].dt.month
            final_df["Anio"]      = final_df["Fecha"].dt.year
            final_df["Mes"]       = final_df["Mes_num"].map(MESES_ES)
            final_df["Mes_label"] = final_df.apply(
                lambda r: "{} {}".format(r["Mes"], int(r["Anio"])), axis=1
            )
            orden = (
                final_df[["Anio", "Mes_num", "Mes_label"]]
                .drop_duplicates()
                .sort_values(["Anio", "Mes_num"])
            )
            meses_ordenados = list(orden["Mes_label"])
            data_por_mes = {}
            for mes in meses_ordenados:
                df_mes = final_df[final_df["Mes_label"] == mes]
                if not df_mes.empty:
                    data_por_mes[mes] = df_mes
        else:
            data_por_mes = {"Sin fecha": final_df}

        st.success("{} mes(es) | {} banco(s)".format(len(data_por_mes), len(banco_cols)))

        tabs = st.tabs(list(data_por_mes.keys()) + ["Total"])
        df_todo = pd.concat(list(data_por_mes.values()), ignore_index=True)
        dfs_preview = list(data_por_mes.items()) + [("Total", df_todo)]
        for tab, (mes_label, df_mes) in zip(tabs, dfs_preview):
            with tab:
                pivot = build_pivot_concepto(df_mes, banco_cols)
                st.dataframe(pivot, use_container_width=True, hide_index=True)

        excel_buf = build_excel(data_por_mes, banco_cols)
        st.download_button(
            "Descargar Consolidado Bancos",
            excel_buf,
            "Consolidado_Bancos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error("Error: {}".format(str(e)))
        st.exception(e)