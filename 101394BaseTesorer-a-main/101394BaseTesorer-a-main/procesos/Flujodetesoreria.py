# -*- coding: utf-8 -*-
import streamlit as st
import zipfile
import io
import pandas as pd
import re
import unicodedata
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOMBRES_SALIDA = ["consolidadobancos", "flujotesoreria", "seguimientorecaudo"]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

KEYWORDS_CATEGORIA = [
    "categoria", "tipo", "tipo de movimiento", "tipo movimiento",
    "tipo transaccion", "clase", "naturaleza",
]
KEYWORDS_CONCEPTO = [
    "concepto", "descripcion", "description", "detalle",
    "referencia", "movimiento", "glosa", "observacion",
]
KEYWORDS_VALOR = [
    "vlr flujo", "valor dc", "valor d/c", "vlr", "valor",
    "valor de la compra", "importe", "monto", "debito credito",
]
KEYWORDS_FECHA = [
    "fecha", "date", "fecha movimiento", "fecha valor",
    "fecha transaccion", "fecha operacion", "fec",
]

VALORES_INGRESO = {"ingreso", "ingresos", "credito", "creditos", "entrada", "entradas"}
VALORES_EGRESO = {
    "egreso", "egresos", "gasto", "gastos", "salida", "salidas",
    "debito", "debitos", "pago", "pagos",
}


def normalize(text):
    text = str(text).lower().strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", text)


def es_archivo_salida(nombre):
    n = normalize(nombre)
    return any(s in n for s in NOMBRES_SALIDA)


def find_column(columns, keywords):
    normalized_cols = {normalize(col): col for col in columns}
    for key in keywords:
        key_norm = normalize(key)
        for col_norm, original_col in normalized_cols.items():
            if key_norm in col_norm or col_norm in key_norm:
                return original_col
    return None


def clean_money(series):
    return (
        series.astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0)
    )


def is_header_row(row):
    text_cells = [str(x).strip() for x in row if isinstance(x, str) and len(str(x).strip()) > 0]
    short_texts = [t for t in text_cells if len(t) < 30]
    return len(short_texts) >= 4


def read_real_excel(file):
    df_raw = pd.read_excel(file, header=None)
    header_row = None
    for i, row in df_raw.iterrows():
        if is_header_row(row):
            header_row = i
            break
    if header_row is None:
        header_row = 0
    return pd.read_excel(file, header=header_row), header_row


def cargar_excels_desde_uploads(uploads):
    result = []
    for upload in uploads:
        name = upload.name
        if es_archivo_salida(name):
            continue
        data = upload.read()
        if name.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for fname in z.namelist():
                    if fname.lower().endswith((".xlsx", ".xls")) and not fname.startswith("__"):
                        fname_base = fname.split("/")[-1]
                        if not es_archivo_salida(fname_base):
                            result.append((fname_base, z.read(fname)))
        elif name.lower().endswith((".xlsx", ".xls")):
            result.append((name, data))
    return result


def standardize_df(df, filename):
    cols = df.columns.tolist()
    n = len(df)

    categoria_col = find_column(cols, KEYWORDS_CATEGORIA)
    concepto_col = find_column(cols, KEYWORDS_CONCEPTO)
    valor_col = find_column(cols, KEYWORDS_VALOR)
    fecha_col = find_column(cols, KEYWORDS_FECHA)

    categoria = df[categoria_col].fillna("Sin categoria") if categoria_col else pd.Series(["Sin categoria"] * n)
    concepto = df[concepto_col].fillna("Sin concepto") if concepto_col else pd.Series(["Sin concepto"] * n)
    total = clean_money(df[valor_col]) if valor_col else pd.Series([0.0] * n)

    if fecha_col:
        parsed = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=True)
        if hasattr(parsed.dt, "tz") and parsed.dt.tz is not None:
            parsed = parsed.dt.tz_convert(None)
        fecha = parsed.dt.normalize()
    else:
        fecha = pd.Series([pd.NaT] * n)

    banco = str(filename).replace(".xlsx", "").replace(".xls", "")

    return pd.DataFrame({
        "Categoria": categoria.astype(str).values,
        "Concepto": concepto.astype(str).values,
        "Banco": banco,
        "Fecha": fecha.values,
        "Valor": total.values,
    }), {
        "categoria_col": categoria_col,
        "concepto_col": concepto_col,
        "valor_col": valor_col,
        "fecha_col": fecha_col,
    }


def es_ingreso(cat_str):
    return normalize(cat_str) in VALORES_INGRESO


def es_egreso(cat_str):
    return normalize(cat_str) in VALORES_EGRESO


def build_excel_flujo(final_pivot, existing_mes_cols):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo de Tesoreria"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", fgColor="BDD7EE")
    total_font = Font(bold=True, size=12)
    ingreso_fill = PatternFill("solid", fgColor="E2EFDA")
    egreso_fill = PatternFill("solid", fgColor="FCE4D6")
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = "#,##0.00"

    cols = list(final_pivot.columns)

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_idx = 2
    for _, row in final_pivot.iterrows():
        cat_upper = str(row.get("Categoria", "")).strip().upper()
        is_flujo_neto = cat_upper in ("FLUJO NETO", "SALDO ACUMULADO")
        is_ingreso_row = cat_upper == "TOTAL INGRESOS"
        is_egreso_row = cat_upper == "TOTAL EGRESOS"
        is_subtotal = cat_upper in ("TOTAL OTROS",)

        for col_idx, col_name in enumerate(cols, start=1):
            value = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            if is_flujo_neto:
                cell.fill = total_fill
                cell.font = total_font
            elif is_ingreso_row:
                cell.fill = ingreso_fill
                cell.font = Font(bold=True, size=11)
            elif is_egreso_row:
                cell.fill = egreso_fill
                cell.font = Font(bold=True, size=11)
            elif is_subtotal:
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(bold=True, size=11)
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

            if col_name not in ("Categoria", "Concepto") and isinstance(value, (int, float)):
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

        row_idx += 1

    for col_idx in range(1, len(cols) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "C2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run():
    st.title("Flujo de Tesoreria")
    st.markdown(
        "Sube los archivos Excel de los bancos **o un ZIP** que los contenga. "
        "El sistema consolida todos los conceptos por mes."
    )

    uploads = st.file_uploader(
        "Suba los Excel o el ZIP con los archivos",
        type=["xlsx", "xls", "zip"],
        accept_multiple_files=True,
        key="ft_files",
    )

    if not uploads:
        return

    try:
        archivos = cargar_excels_desde_uploads(uploads)
        omitidos = [u.name for u in uploads if es_archivo_salida(u.name)]
        if omitidos:
            st.info("Archivos omitidos (son reportes generados): {}".format(", ".join(omitidos)))

        if not archivos:
            st.warning("No se encontraron archivos de bancos validos.")
            return

        st.info("Archivos de banco a procesar: {}".format(len(archivos)))

        all_data = []
        report = []

        for nombre, data in archivos:
            try:
                df, _ = read_real_excel(io.BytesIO(data))
                clean_df, detected = standardize_df(df, nombre)
                fechas_ok = int(clean_df["Fecha"].notna().sum())
                categorias_unicas = clean_df["Categoria"].unique().tolist()[:10]
                report.append({
                    "Archivo": nombre,
                    "Filas": len(clean_df),
                    "Col Categoria": detected["categoria_col"] or "NO DETECTADA",
                    "Col Valor": detected["valor_col"] or "NO DETECTADA",
                    "Col Fecha": detected["fecha_col"] or "NO DETECTADA",
                    "Filas con fecha": fechas_ok,
                    "Categorias encontradas": str(categorias_unicas),
                    "Estado": "OK",
                })
                all_data.append(clean_df)
            except Exception as e:
                report.append({
                    "Archivo": nombre,
                    "Filas": 0,
                    "Col Categoria": "ERROR",
                    "Col Valor": "ERROR",
                    "Col Fecha": "ERROR",
                    "Filas con fecha": 0,
                    "Categorias encontradas": "",
                    "Estado": "Error: {}".format(str(e)),
                })

        with st.expander("Diagnostico de lectura por archivo", expanded=True):
            st.dataframe(pd.DataFrame(report), use_container_width=True)
            st.markdown(
                "Si 'Col Categoria' dice **NO DETECTADA**, el sistema no puede "
                "separar Ingresos de Egresos. Comparta una captura de las columnas "
                "de su Excel para corregirlo."
            )

        if not all_data:
            st.error("Ningun archivo aporto datos utiles.")
            return

        final_df = pd.concat(all_data, ignore_index=True)
        df_con_fecha = final_df[final_df["Fecha"].notna()].copy()
        n_sin_fecha = int(final_df["Fecha"].isna().sum())

        if n_sin_fecha > 0:
            st.warning("{} filas sin fecha valida seran ignoradas.".format(n_sin_fecha))

        if df_con_fecha.empty:
            st.error(
                "No se encontro columna de fecha en ninguno de los archivos. "
                "El Flujo de Tesoreria requiere fechas para agrupar por mes."
            )
            return

        cats_unicas = sorted(df_con_fecha["Categoria"].unique().tolist())
        with st.expander("Categorias unicas encontradas en todos los archivos"):
            st.write(cats_unicas)
            st.markdown(
                "El sistema clasifica como **Ingresos**: ingreso, ingresos, credito, entrada. "
                "Como **Egresos**: egreso, egresos, gasto, gastos, salida, debito, pago."
            )

        df_con_fecha["Mes_num"] = df_con_fecha["Fecha"].dt.month
        df_con_fecha["Anio"] = df_con_fecha["Fecha"].dt.year
        df_con_fecha["Mes"] = df_con_fecha["Mes_num"].map(MESES_ES)
        df_con_fecha["Mes_label"] = df_con_fecha.apply(
            lambda r: "{} {}".format(r["Mes"], int(r["Anio"])), axis=1
        )

        meses_presentes = (
            df_con_fecha[["Anio", "Mes_num", "Mes_label"]]
            .drop_duplicates()
            .sort_values(["Anio", "Mes_num"])
        )
        meses_cols = list(meses_presentes["Mes_label"])

        grouped = (
            df_con_fecha
            .groupby(["Categoria", "Concepto", "Mes_label"])["Valor"]
            .sum()
            .reset_index()
        )

        if grouped.empty:
            st.error("No hay datos para agrupar tras filtrar por fecha.")
            return

        pivot_df = grouped.pivot_table(
            index=["Categoria", "Concepto"],
            columns="Mes_label",
            values="Valor",
            fill_value=0,
        ).reset_index()
        pivot_df.columns.name = None

        existing_mes_cols = [m for m in meses_cols if m in pivot_df.columns]
        for m in meses_cols:
            if m not in pivot_df.columns:
                pivot_df[m] = 0.0
        pivot_df = pivot_df[["Categoria", "Concepto"] + existing_mes_cols]
        pivot_df["Total"] = pivot_df[existing_mes_cols].sum(axis=1)

        mask_ingreso = pivot_df["Categoria"].apply(es_ingreso)
        mask_egreso = pivot_df["Categoria"].apply(es_egreso)

        ingresos_df = pivot_df[mask_ingreso].copy()
        egresos_df = pivot_df[mask_egreso].copy()
        otros_df = pivot_df[~mask_ingreso & ~mask_egreso].copy()

        st.write(
            "**Clasificacion:** Ingresos: {} filas | Egresos: {} filas | Otros: {} filas".format(
                len(ingresos_df), len(egresos_df), len(otros_df)
            )
        )

        num_cols = existing_mes_cols + ["Total"]

        def totals_row(df, label_cat, label_conc=""):
            row = {c: float(df[c].sum()) for c in num_cols}
            row["Categoria"] = label_cat
            row["Concepto"] = label_conc
            return pd.DataFrame([row])

        sections = []

        if not ingresos_df.empty:
            sections.append(ingresos_df)
            sections.append(totals_row(ingresos_df, "TOTAL INGRESOS"))

        if not egresos_df.empty:
            sections.append(egresos_df)
            sections.append(totals_row(egresos_df, "TOTAL EGRESOS"))

        if not otros_df.empty:
            sections.append(otros_df)
            sections.append(totals_row(otros_df, "OTROS", "TOTAL OTROS"))

        if not ingresos_df.empty or not egresos_df.empty:
            flujo_row = {}
            for c in num_cols:
                ing = float(ingresos_df[c].sum()) if not ingresos_df.empty else 0.0
                egr = float(egresos_df[c].sum()) if not egresos_df.empty else 0.0
                flujo_row[c] = ing - egr
            flujo_row["Categoria"] = "FLUJO NETO"
            flujo_row["Concepto"] = "Ingresos - Egresos"
            sections.append(pd.DataFrame([flujo_row]))

            saldo_row = {"Categoria": "SALDO ACUMULADO", "Concepto": "Acumulado del periodo"}
            acum = 0.0
            for c in existing_mes_cols:
                acum += flujo_row[c]
                saldo_row[c] = acum
            saldo_row["Total"] = flujo_row["Total"]
            sections.append(pd.DataFrame([saldo_row]))

        if not sections:
            st.error(
                "No se generaron secciones. Revisa el diagnostico de arriba "
                "para ver las categorias encontradas."
            )
            return

        final_pivot = pd.concat(sections, ignore_index=True)
        col_order = ["Categoria", "Concepto"] + existing_mes_cols + ["Total"]
        final_pivot = final_pivot[col_order]

        st.success(
            "Flujo de Tesoreria listo: {} mes(es) | {} conceptos".format(
                len(existing_mes_cols), len(pivot_df)
            )
        )

        df_show = final_pivot.copy()
        for c in existing_mes_cols + ["Total"]:
            df_show[c] = df_show[c].apply(
                lambda x: "${:,.2f}".format(x) if isinstance(x, (int, float)) else x
            )
        st.dataframe(df_show, use_container_width=True, hide_index=True)

        excel_buf = build_excel_flujo(final_pivot, existing_mes_cols)
        st.download_button(
            "Descargar Flujo de Tesoreria",
            excel_buf,
            "Flujo_de_Tesoreria.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error("Error inesperado: {}".format(str(e)))
        st.exception(e)